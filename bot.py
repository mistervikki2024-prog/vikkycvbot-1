from flask import Flask
import os
import threading
import json
import time
import telebot
from telebot import types

# 🔹 Flask app
web = Flask(__name__)

@web.route('/')
def home():
    return "Bot is running!"

# 🔹 Config
TOKEN = os.getenv("BOT_TOKEN")
ADMIN_ID = int(os.getenv("ADMIN_ID", "5328734113"))

bot = telebot.TeleBot(TOKEN)



# ============================================================
# 🔹 MAIN MENU — Colored Buttons + Animated Emoji
# ============================================================
def main_menu():
    kb = types.ReplyKeyboardMarkup(resize_keyboard=True)

    kb.row("📄 Text to VCF", "📑 VCF to Text")
    kb.row("🧩 Merge VCF", "🧩 Merge Text")
    kb.row("✂️ Split VCF", "✂️ Split Text")
    kb.row("✏️ VCF Editor", "🔍 Get VCF Details")
    kb.row("👑 My Subscription")

    return kb

# ============================================================
# 🔹 User State
# ============================================================
user_state = {}

# ============================================================
# 🔹 Load / Save Users
# ============================================================
def load_users():
    try:
        with open("users.json", "r") as f:
            return json.load(f)
    except:
        return {}

def save_users(data):
    with open("users.json", "w") as f:
        json.dump(data, f, indent=4)


# ============================================================
# 🔹 Progress Bar
# ============================================================
def progress_bar(current, total):
    percent = int((current / total) * 100) if total else 0
    filled = int(percent / 5)
    bar = "█" * filled + "░" * (20 - filled)
    return f"{bar} {percent}%"

# ============================================================
# 🔹 /start
# ============================================================
@bot.message_handler(commands=["start"])
def start(message):
    users = load_users()
    uid = str(message.from_user.id)

    if uid not in users:
        users[uid] = {"premium": False}
        save_users(users)

    bot.send_message(
        message.chat.id,
        (
            "🔥 *WELCOME TO VCF TOOL BOT* 🔥\n"
        ),
        parse_mode="Markdown",
        reply_markup=main_menu()
    )

# ============================================================
# 🔹 TEXT HANDLER
# ============================================================
@bot.message_handler(func=lambda m: True, content_types=["text"])
def handle_text(message):
    user_id = message.from_user.id
    text = message.text.strip()
    state = user_state.get(user_id)

    # ── MENU BUTTONS ──────────────────────────────────────────

    if text == "Text to VCF":
        start_txt_to_vcf(message, user_id)
        return

    if text == "VCF to Text":
        start_vcf_to_txt(message, user_id)
        return

    if text == "Manual VCF":
        start_merge_vcf(message, user_id)
        return

    if text == "Manual Text":
        bot.send_message(message.chat.id, "✂️ Use *Text to VCF* with a contact limit.", parse_mode="Markdown")
        return

    if text == "Admin/Navy VCF":
        if user_id == ADMIN_ID:
            users = load_users()
            bot.send_message(message.chat.id, f"👥 *Total Users:* {len(users)}", parse_mode="Markdown")
        else:
            bot.send_message(message.chat.id, "❌ Not allowed.")
        return

    if text == "⚙️ My Subscription":
        bot.send_message(message.chat.id, "⚙️ *My Subscription*\n\nContact admin for premium.", parse_mode="Markdown")
        return

    if text == "🔍 Get Name":
        bot.send_message(message.chat.id, "🔍 *Get Name* feature coming soon!", parse_mode="Markdown")
        return

    if text == "✏️ VCF Editor":
        bot.send_message(message.chat.id, "✏️ *VCF Editor* feature coming soon!", parse_mode="Markdown")
        return

    if text == "📑 Merge TEXT":
        bot.send_message(message.chat.id, "📑 *Merge TEXT* feature coming soon!", parse_mode="Markdown")
        return

    # ── STATE FLOWS ───────────────────────────────────────────

    if not state:
        bot.send_message(message.chat.id, "⚠️ Please select an option from menu first.", reply_markup=main_menu())
        return

    mode = state.get("mode")

    # ── TEXT TO VCF ────────────────────────────────────────────
    if mode == "collect":
        if text == "/done":
            if not state["numbers"]:
                bot.send_message(message.chat.id, "❌ No contacts added yet.")
                return
            state["mode"] = "ask_name"
            bot.send_message(message.chat.id, "1️⃣ *VCF File Name?*\n_(Example: Hongkong)_", parse_mode="Markdown")
        else:
            for n in text.split():
                n = n.replace(" ", "").replace("-", "").replace("+", "")
                if n.isdigit() and len(n) >= 8:
                    state["numbers"].append(n)
        return

    if mode == "ask_name":
        state["file_name"] = text
        state["mode"] = "ask_prefix"
        bot.send_message(message.chat.id, "2️⃣ *Contact Name Prefix?*\n_(Example: Vikky Boss)_", parse_mode="Markdown")
        return

    if mode == "ask_prefix":
        state["prefix"] = text
        state["mode"] = "ask_start_vcf"
        bot.send_message(message.chat.id, "3️⃣ *VCF File Starting Number?*\n_(Example: 1)_", parse_mode="Markdown")
        return

    if mode == "ask_start_vcf":
        try:
            state["vcf_start"] = int(text)
        except:
            bot.send_message(message.chat.id, "❌ Enter a valid number.")
            return
        state["mode"] = "ask_contact_start"
        bot.send_message(message.chat.id, "4️⃣ *Contact Starting Number?*\n_(Example: 1)_", parse_mode="Markdown")
        return

    if mode == "ask_contact_start":
        try:
            state["contact_start"] = int(text)
        except:
            bot.send_message(message.chat.id, "❌ Enter a valid number.")
            return
        state["mode"] = "ask_limit"
        bot.send_message(message.chat.id, "5️⃣ *Contacts per VCF file?*\n_(Example: 50)_", parse_mode="Markdown")
        return

    if mode == "ask_limit":
        try:
            limit = int(text)
        except:
            bot.send_message(message.chat.id, "❌ Enter a valid number.")
            return

        numbers = state["numbers"]
        bot.send_message(
            message.chat.id,
            f"🚀 *Generating VCF Files*\n━━━━━━━━━━━━━━━\n"
            f"📊 Total Contacts: {len(numbers)}\n⚡ Status: Processing...",
            parse_mode="Markdown"
        )

        chunks = [numbers[i:i+limit] for i in range(0, len(numbers), limit)]
        contact_counter = state["contact_start"]

        for idx, chunk in enumerate(chunks):
            vcf_data = ""
            for num in chunk:
                vcf_data += f"BEGIN:VCARD\nVERSION:3.0\nFN:{state['prefix']} {contact_counter}\nTEL;TYPE=CELL:{num}\nEND:VCARD\n"
                contact_counter += 1

            filename = f"{state['file_name']}{state['vcf_start'] + idx}.vcf"
            with open(filename, "w") as f:
                f.write(vcf_data)

            with open(filename, "rb") as f:
                bot.send_document(message.chat.id, f)
            os.remove(filename)

        bot.send_message(message.chat.id, "✅ *VCF Generation Completed Successfully!* 🎉", parse_mode="Markdown")
        user_state.pop(user_id, None)
        return

    # ── VCF TO TXT ─────────────────────────────────────────────
    if mode == "vcf_to_txt":
        if text == "/done":
            state["animating"] = False
            time.sleep(0.6)

            final_text = (
                f"📄 *Final Result*\n━━━━━━━━━━━━━━━\n"
                f"📁 Files Processed: {state.get('files', 0)}\n"
                f"📊 Total Extracted: {len(state['numbers'])}\n"
                f"✅ Finished!"
            )

            if state.get("msg_id"):
                try:
                    bot.edit_message_text(
                        final_text,
                        message.chat.id,
                        state["msg_id"],
                        parse_mode="Markdown"
                    )
                except:
                    pass
            else:
                bot.send_message(message.chat.id, final_text, parse_mode="Markdown")

            state["step"] = "ask_name"
            bot.send_message(message.chat.id, "📝 *Enter the name for your .txt file:*\n_(Example: ExtractedList)_", parse_mode="Markdown")
        return

    if mode == "vcf_to_txt" and state.get("step") == "ask_name":
        filename = f"{text}.txt"
        with open(filename, "w") as f:
            f.write("\n".join(state["numbers"]))

        with open(filename, "rb") as f:
            bot.send_document(message.chat.id, f)
        os.remove(filename)

        bot.send_message(message.chat.id, "✅ *Extraction Completed Successfully!* 🎉", parse_mode="Markdown")
        user_state.pop(user_id, None)
        return

    # ── MERGE VCF ──────────────────────────────────────────────
    if mode == "merge_vcf":
        step = state.get("step")

        if step == "ask_filename":
            state["filename"] = text
            state["step"] = "ask_prefix"
            bot.send_message(message.chat.id, "✏️ *Enter contact name prefix:*", parse_mode="Markdown")
            return

        if step == "ask_prefix":
            state["prefix"] = text
            state["step"] = "collecting"
            state["all_numbers"] = []
            bot.send_message(message.chat.id, "📤 *Send all VCF files, then type* `DONE`", parse_mode="Markdown")
            return

        if text.upper() == "DONE" and step == "collecting":
            numbers = list(set(state.get("all_numbers", [])))

            if not numbers:
                bot.send_message(message.chat.id, "❌ No data found.")
                return

            vcf_data = ""
            for i, num in enumerate(numbers):
                vcf_data += f"BEGIN:VCARD\nVERSION:3.0\nFN:{state['prefix']} {i+1}\nTEL;TYPE=CELL:{num}\nEND:VCARD\n"

            filename = f"{state['filename']}.vcf"
            with open(filename, "w") as f:
                f.write(vcf_data)

            with open(filename, "rb") as f:
                bot.send_document(message.chat.id, f)
            os.remove(filename)

            user_state.pop(user_id, None)
            bot.send_message(message.chat.id, "✅ *All VCF files merged!* 🎉", parse_mode="Markdown")
            return

# ============================================================
# 🔹 Helper: Start Modes
# ============================================================
def start_txt_to_vcf(message, user_id):
    user_state[user_id] = {
        "mode": "collect",
        "numbers": [],
        "files": 0,
        "start_time": time.time()
    }
    bot.send_message(
        message.chat.id,
        "📥 *Send Contacts*\n═══════════════\n📂 Numbers / .txt / .xlsx\n\n✅ *Finish* → Type `/done`",
        parse_mode="Markdown"
    )

def start_vcf_to_txt(message, user_id):
    user_state[user_id] = {
        "mode": "vcf_to_txt",
        "numbers": [],
        "files": 0,
        "msg_id": None,
        "start_time": time.time(),
        "total_lines": 0,
        "processed_lines": 0,
        "animating": False
    }
    bot.send_message(
        message.chat.id,
        "📤 *Upload VCF Files*\n━━━━━━━━━━━━━━━\n📁 Send one or multiple `.vcf` files\n\n✅ *Finish* → Type `/done`",
        parse_mode="Markdown"
    )

def start_merge_vcf(message, user_id):
    user_state[user_id] = {
        "mode": "merge_vcf",
        "step": "ask_filename"
    }
    bot.send_message(message.chat.id, "📝 *Enter output VCF file name:*", parse_mode="Markdown")

# ============================================================
# 🔹 Animate Progress
# ============================================================
def animate_progress(chat_id, msg_id, state):
    last_done = 0
    last_time = time.time()

    while state.get("animating"):
        time.sleep(0.5)
        total = max(state.get("total_lines", 1), 1)
        done = state.get("processed_lines", 0)

        now = time.time()
        speed = (done - last_done) / (now - last_time) if (now - last_time) > 0 else 0
        last_done = done
        last_time = now

        percent = min(int((done / total) * 100), 100)
        filled = int(percent / 5)
        bar = "█" * filled + "░" * (20 - filled)

        text_msg = (
            f"🚀 *VCF SCANNING*\n"
            f"━━━━━━━━━━━━━━━\n\n"
            f"📁 Files: {state.get('files', 0)}\n"
            f"📊 Extracted: {len(state.get('numbers', []))}\n\n"
            f"📈 Progress: `{bar} {percent}%`\n\n"
            f"⚡ Speed: {speed:.0f} lines/sec\n"
            f"🔄 {done}/{total} lines"
        )

        try:
            bot.edit_message_text(text_msg, chat_id, msg_id, parse_mode="Markdown")
        except:
            pass

# ============================================================
# 🔹 Process VCF File
# ============================================================
def process_vcf_file(path, state):
    with open(path, encoding="utf-8", errors="ignore") as f:
        for line in f:
            state["total_lines"] += 1
            line = line.strip()
            if "TEL" in line.upper():
                num = line.split(":")[-1].strip()
                num = num.replace(" ", "").replace("-", "").replace("+", "")
                if num.isdigit() and len(num) >= 8:
                    state["numbers"].append(num)
            state["processed_lines"] += 1
    try:
        os.remove(path)
    except:
        pass

# ============================================================
# 🔹 FILE HANDLER
# ============================================================
@bot.message_handler(content_types=["document"])
def handle_files(message):
    user_id = message.from_user.id
    state = user_state.get(user_id)
    doc = message.document
    filename = doc.file_name.lower()

    if not state:
        bot.send_message(message.chat.id, "⚠️ Please select an option from menu first.")
        return

    file_info = bot.get_file(doc.file_id)
    path = f"{user_id}_{filename}"

    downloaded = bot.download_file(file_info.file_path)
    with open(path, "wb") as f:
        f.write(downloaded)

    mode = state.get("mode")

    # ── TXT file for TEXT TO VCF ──────────────────────────────
    if filename.endswith(".txt") and mode == "collect":
        with open(path) as f:
            for line in f:
                num = line.strip().replace(" ", "").replace("-", "").replace("+", "")
                if num.isdigit() and len(num) >= 8:
                    state["numbers"].append(num)
        os.remove(path)
        bot.send_message(
            message.chat.id,
            f"📥 *Contacts Added:* {len(state['numbers'])}\n✅ Send more or type `/done`",
            parse_mode="Markdown"
        )
        return

    # ── XLSX file for TEXT TO VCF ─────────────────────────────
    if filename.endswith(".xlsx") and mode == "collect":
        try:
            from openpyxl import load_workbook
            wb = load_workbook(path)
            sheet = wb.active
            for row in sheet.iter_rows(values_only=True):
                for cell in row:
                    if cell:
                        num = str(cell).strip().replace(" ", "").replace("-", "").replace("+", "")
                        if num.isdigit() and len(num) >= 8:
                            state["numbers"].append(num)
        except Exception as e:
            bot.send_message(message.chat.id, f"❌ XLSX error: {e}")
        os.remove(path)
        bot.send_message(
            message.chat.id,
            f"📥 *Contacts Added:* {len(state['numbers'])}\n✅ Send more or type `/done`",
            parse_mode="Markdown"
        )
        return

    # ── VCF file for VCF TO TXT ───────────────────────────────
    if filename.endswith(".vcf") and mode == "vcf_to_txt":
        state["files"] = state.get("files", 0) + 1

        if not state.get("msg_id"):
            msg = bot.send_message(message.chat.id, "📄 *Starting scan...*", parse_mode="Markdown")
            state["msg_id"] = msg.message_id
            state["animating"] = True
            state["total_lines"] = 0
            state["processed_lines"] = 0

            threading.Thread(
                target=animate_progress,
                args=(message.chat.id, msg.message_id, state),
                daemon=True
            ).start()

        threading.Thread(
            target=process_vcf_file,
            args=(path, state),
            daemon=True
        ).start()
        return

    # ── VCF file for MERGE VCF ────────────────────────────────
    if filename.endswith(".vcf") and mode == "merge_vcf":
        if "all_numbers" not in state:
            state["all_numbers"] = []

        with open(path, encoding="utf-8", errors="ignore") as f:
            for line in f:
                if "TEL" in line.upper():
                    num = line.split(":")[-1].strip()
                    num = num.replace(" ", "").replace("-", "").replace("+", "")
                    if num.isdigit() and len(num) >= 8:
                        state["all_numbers"].append(num)

        os.remove(path)
        bot.send_message(message.chat.id, f"✅ File added. Send more or type `DONE`", parse_mode="Markdown")
        return

    # ── Invalid ───────────────────────────────────────────────
    try:
        os.remove(path)
    except:
        pass
    bot.send_message(message.chat.id, "❌ Invalid file type for current mode.")

# ============================================================
# 🔹 /help
# ============================================================
@bot.message_handler(commands=["help"])
def help_cmd(message):
    bot.send_message(
        message.chat.id,
        (
            "📖 *VCF Tool Bot — Help*\n"
            "━━━━━━━━━━━━━━━\n\n"
            "📁 *Text to VCF* — Convert numbers to VCF\n"
            "📄 *VCF to Text* — Extract numbers from VCF\n"
            "👑 *Admin/Navy VCF* — Admin stats\n"
            "🔄 *Merge VCF* — Merge multiple VCF files\n"
            "✂️ *Split VCF* — Split contacts into files\n"
            "⚙️ *My Subscription* — Check your plan\n\n"
            "📞 *Support:* @Vikky_IND"
        ),
        parse_mode="Markdown"
    )

# ============================================================
# 🔹 Run Bot
# ============================================================
def run_bot():
    print("✅ Bot starting with pyTelegramBotAPI...")
    if not TOKEN:
        print("❌ BOT_TOKEN missing!")
        return
    bot.infinity_polling(timeout=10, long_polling_timeout=5)

threading.Thread(target=run_bot, daemon=True).start()

if __name__ == "__main__":
    port = int(os.getenv("PORT", 5000))
    web.run(host="0.0.0.0", port=port)